"""
Excel Writer Module (Dynamic Columns)
Automatically creates columns based on extracted OCR fields
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from datetime import datetime


class ExcelWriter:
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.ensure_file_exists()

    def ensure_file_exists(self):
        path = Path(self.filepath)
        path.parent.mkdir(parents=True, exist_ok=True)

        if not path.exists():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Scanned Data"
            wb.save(self.filepath)

    def _apply_header_style(self, cell):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="366092")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

    def _apply_cell_style(self, cell, center=False):
        cell.alignment = Alignment(
            horizontal="center" if center else "left",
            vertical="center"
        )
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

    def append_row(self, fields: dict):
        """
        Append OCR extracted data dynamically
        """
        wb = openpyxl.load_workbook(self.filepath)
        ws = wb.active

        # Existing headers
        headers = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []

        # Add Timestamp automatically
        fields = dict(fields)
        fields["Timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Create headers if empty
        if not headers or headers == [None]:
            for col, header in enumerate(fields.keys(), start=1):
                cell = ws.cell(row=1, column=col, value=header)
                self._apply_header_style(cell)
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

            headers = list(fields.keys())

        # Add new headers if new fields appear
        for key in fields.keys():
            if key not in headers:
                headers.append(key)
                col = len(headers)
                cell = ws.cell(row=1, column=col, value=key)
                self._apply_header_style(cell)
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

        # Append data row
        row = ws.max_row + 1
        for col, header in enumerate(headers, start=1):
            value = fields.get(header, "")
            cell = ws.cell(row=row, column=col, value=value)
            self._apply_cell_style(cell, center=("Contact" in header or "Marks" in header))

        ws.row_dimensions[row].height = 20
        wb.save(self.filepath)

        print(f"✓ Data saved to Excel (row {row})")

    def get_row_count(self):
        wb = openpyxl.load_workbook(self.filepath)
        ws = wb.active
        return ws.max_row - 1
