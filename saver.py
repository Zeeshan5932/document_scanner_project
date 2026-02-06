"""
Saver Module
Saves extracted data in multiple formats with automatic timestamp.
Supported formats: Excel (.xlsx), JSON (.json), CSV (.csv)
Only saves after explicit user confirmation.
"""

import json
import csv
import os
from datetime import datetime
from pathlib import Path
from typing import Dict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class DataSaver:
    """Saves extracted form data in multiple formats with timestamps."""
    
    def __init__(self, output_dir: str = "output"):
        """
        Initialize saver with output directory.
        
        Args:
            output_dir (str): Directory to save files (default: "output")
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.timestamp = datetime.now()
        print(f"✓ Output directory: {self.output_dir}")
    
    def save(self, extracted_data: Dict, filename: str, format_type: str) -> bool:
        """
        Save extracted data in requested format.
        
        Args:
            extracted_data (dict): Dictionary of extracted fields
            filename (str): Base filename (without extension)
            format_type (str): Format type: 'excel', 'json', or 'csv'
        
        Returns:
            bool: True if save successful, False otherwise
        """
        format_type = format_type.lower().strip()
        
        if format_type == "excel":
            return self._save_excel(extracted_data, filename)
        elif format_type == "json":
            return self._save_json(extracted_data, filename)
        elif format_type == "csv":
            return self._save_csv(extracted_data, filename)
        else:
            print(f"❌ ERROR: Unknown format: {format_type}")
            return False
    
    def _save_excel(self, extracted_data: Dict, filename: str) -> bool:
        """
        Save data as Excel file (.xlsx).
        
        Creates spreadsheet with:
        - Metadata (timestamp, source)
        - Header row with field names
        - Data row with values and confidence scores
        - Formatted cells
        
        Args:
            extracted_data (dict): Extracted fields
            filename (str): Base filename
        
        Returns:
            bool: Success status
        """
        if not OPENPYXL_AVAILABLE:
            print("❌ ERROR: openpyxl not installed")
            print("   Install with: pip install openpyxl")
            return False
        
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Scanned Data"
            
            # Add metadata
            sheet["A1"] = "Scanned Form Data"
            sheet["A1"].font = Font(bold=True, size=14)
            sheet.merge_cells("A1:D1")
            
            sheet["A2"] = f"Timestamp: {self.timestamp.strftime('%Y-%m-%d %H:%M:%S')}"
            sheet["A2"].font = Font(italic=True, size=10)
            sheet.merge_cells("A2:D2")
            
            sheet["A3"] = ""  # Blank row
            
            # Create header row with field names
            field_names = sorted(extracted_data.keys())
            for col_idx, field_name in enumerate(field_names, 1):
                cell = sheet.cell(row=4, column=col_idx)
                cell.value = field_name
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Add data row
            for col_idx, field_name in enumerate(field_names, 1):
                cell = sheet.cell(row=5, column=col_idx)
                value = extracted_data[field_name].get("value", "")
                confidence = extracted_data[field_name].get("confidence", 0)
                
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # Add confidence row
            sheet.cell(row=6, column=1).value = "Confidence (%)"
            sheet.cell(row=6, column=1).font = Font(italic=True)
            for col_idx, field_name in enumerate(field_names, 1):
                cell = sheet.cell(row=6, column=col_idx)
                confidence = extracted_data[field_name].get("confidence", 0)
                cell.value = confidence
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for col_idx, field_name in enumerate(field_names, 1):
                column_letter = sheet.cell(row=4, column=col_idx).column_letter
                sheet.column_dimensions[column_letter].width = max(18, len(field_name) + 3)
            
            # Save file
            filepath = self.output_dir / f"{filename}.xlsx"
            workbook.save(filepath)
            
            print(f"✓ Excel file saved: {filepath}")
            return True
            
        except Exception as e:
            print(f"❌ ERROR saving Excel file: {e}")
            return False
    
    def _save_json(self, extracted_data: Dict, filename: str) -> bool:
        """
        Save data as JSON file (.json).
        
        JSON structure:
        {
            "timestamp": "2026-02-04 14:30:45",
            "extracted_fields": {
                "field_name": {
                    "value": "...",
                    "confidence": 85
                },
                ...
            }
        }
        
        Args:
            extracted_data (dict): Extracted fields
            filename (str): Base filename
        
        Returns:
            bool: Success status
        """
        try:
            data = {
                "timestamp": self.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                "extracted_fields": extracted_data,
            }
            
            filepath = self.output_dir / f"{filename}.json"
            
            with open(filepath, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            
            print(f"✓ JSON file saved: {filepath}")
            return True
            
        except Exception as e:
            print(f"❌ ERROR saving JSON file: {e}")
            return False
    
    def _save_csv(self, extracted_data: Dict, filename: str) -> bool:
        """
        Save data as CSV file (.csv).
        
        CSV structure:
        - Row 1: Timestamp
        - Row 2: Field names (header)
        - Row 3: Extracted values
        - Row 4: Confidence scores
        
        Args:
            extracted_data (dict): Extracted fields
            filename (str): Base filename
        
        Returns:
            bool: Success status
        """
        try:
            filepath = self.output_dir / f"{filename}.csv"
            
            field_names = sorted(extracted_data.keys())
            
            # Calculate average confidence
            confidence_values = [
                extracted_data[fn].get("confidence", 0) 
                for fn in field_names
            ]
            avg_confidence = sum(confidence_values) / len(confidence_values) if confidence_values else 0
            
            with open(filepath, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                
                # Write timestamp
                writer.writerow(["Timestamp", self.timestamp.strftime("%Y-%m-%d %H:%M:%S")])
                writer.writerow([])  # Blank row
                
                # Write header
                writer.writerow(["Field"] + field_names)
                
                # Write values
                value_row = ["Value"]
                for field_name in field_names:
                    value = extracted_data[field_name].get("value", "")
                    value_row.append(value)
                writer.writerow(value_row)
                
                # Write confidence
                confidence_row = ["Confidence (%)"]
                for field_name in field_names:
                    confidence = extracted_data[field_name].get("confidence", 0)
                    confidence_row.append(str(confidence))
                writer.writerow(confidence_row)
                
                # Write average confidence
                writer.writerow([])  # Blank row
                writer.writerow(["Average Confidence", f"{avg_confidence:.1f}%"])
            
            print(f"✓ CSV file saved: {filepath}")
            return True
            
        except Exception as e:
            print(f"❌ ERROR saving CSV file: {e}")
            return False
    
    def list_output_files(self) -> None:
        """List all files in output directory."""
        if not self.output_dir.exists():
            print(f"⚠ Output directory doesn't exist: {self.output_dir}")
            return
        
        files = list(self.output_dir.glob("*"))
        
        if not files:
            print(f"No files in {self.output_dir}")
            return
        
        print(f"\n📁 Output files ({self.output_dir}):")
        for file in sorted(files):
            size_kb = file.stat().st_size / 1024
            print(f"  • {file.name} ({size_kb:.1f} KB)")
